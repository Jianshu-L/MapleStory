from __future__ import annotations
import random
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Iterable, Optional
import pandas as pd
import numpy as np
from utils import flatten_list

# ---------- Helpers ----------

def clean_id(x) -> str:
    # 去掉首尾空白字符
    if pd.isna(x):
        return ""
    return str(x).strip()

def find_contain_index(data: list, target: str):
    return [i for i,j in enumerate(data) if j in target]

def pop_random(lst: List, *, rng: random.Random) -> Optional[str]:
    if not lst:
        return None
    i = rng.randrange(len(lst))
    lst[i], lst[-1] = lst[-1], lst[i]
    return lst.pop()

def get_key_by_value(key_list, value_list, value):
    return [k for k, v in zip(key_list, value_list) if v == value]

# ---------- Mapping inputs to jobs ----------

@dataclass
class MappingSource:
    # 基础数据源：并行数组或从外部注入的主数据
    weights: List[int]      # 每个id的熟悉程度，越熟悉权重越大
    ids: List[str]          # 主数据中的 id 列表（去重后）
    jobs: List[str]         # 对应职业 e.g. "奶" "拳" "眼" "火" 
    job_types: List[str]    # 对应职业标签，e.g. "近战" "远程"

    def __post_init__(self):
        # 创建完就检查长度是否一致
        self._check_lengths()

    def _check_lengths(self):
        lengths = [len(self.weights), len(self.ids), len(self.jobs), len(self.job_types)]
        if len(set(lengths)) != 1:
            raise ValueError(f"the length of data is different: weights={len(self.weights)}, ids={len(self.ids)}, jobs={len(self.jobs)}, job_types={len(self.job_types)}")

    def sort_by_weight(self, reverse: bool = False):
        """按权重排序，小的在前，大的在后（reverse=True 反转顺序）"""
        sorted_data = sorted(zip(self.weights, self.ids, self.jobs, self.job_types),
                             key=lambda x: x[0], reverse=reverse)
        self.weights, self.ids, self.jobs, self.job_types = map(list, zip(*sorted_data))

    def get_weight_by_id(self, target_id: str) -> int:
        """根据 id 返回对应的 weight 值"""
        try:
            index = self.ids.index(target_id)
            return self.weights[index]
        except ValueError:
            raise KeyError(f"id '{target_id}' not found in ids")

    def get_jtype_by_id(self, target_id: str) -> int:
        """根据 id 返回对应的 weight 值"""
        try:
            index = self.ids.index(target_id)
            return self.job_types[index]
        except ValueError:
            raise KeyError(f"id '{target_id}' not found in ids")

def build_index(repo_path: str="repo.csv") -> Tuple[Dict[str, Tuple[str, str, int]], MappingSource]:
    """
    将主数据 ids → (job, job_type) 的唯一映射构建出来。
    """
    df = pd.read_csv(repo_path)
    ms = MappingSource(
        weights=df.index.tolist(),
        ids=df.id.tolist(),
        jobs=df.job.tolist(),
        job_types=df.job_type.tolist(),
    )

    results = {}
    for i, pid in enumerate(ms.ids):
        pid_clean = clean_id(pid)
        if not pid_clean:
            continue
        if pid_clean in results:
            # 如有重复，raise warning and pass
            Warning(f"Dulicate Id {pid_clean} in repo, ignore")
            continue

        job = ms.jobs[i]
        job_type = ms.job_types[i]
        weights = ms.weights[i]
        results[pid_clean] = (job, job_type, weights)
    return results, ms

# ---------- Reporting ----------

@dataclass
class GroupReport:
    warnings: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    unmapped_ids: List[str] = field(default_factory=list)   # CSV 中找不到映射的 ID
    invalid_ids: List[str|tuple] = field(default_factory=list)    # 空/非法/重复 ID
    leftover_buckets: Dict[str, List[str]] = field(default_factory=dict)  # 分组后仍剩余的人
    grouped: List[List[str]] = field(default_factory=list)  # 各队成员 ID
    grouped_jobs: List[List[str]] = field(default_factory=list)  # 各队成员的 job_type
    ok: bool = True

    def add_error(self, msg: str):
        self.errors.append(msg)
        self.ok = False

    def add_warning(self, msg: str):
        self.warnings.append(msg)

# ---------- Core pipeline ----------
def handle_find_contain_index_bug(pid: str, repo_id: List[str], index_list: List[int], report: GroupReport):
    possible_ids = [repo_id[i] for i in index_list]
    sorted_possible_ids = sorted(possible_ids, key=len, reverse=True)
    report.invalid_ids.append((pid, sorted_possible_ids))
    report.add_warning(f"duplicated matching in Repo: {pid} {sorted_possible_ids}")

    argmax = np.argmax([len(w) for w in possible_ids])
    index_list[0] = index_list[argmax]
    return index_list

import re
def map_today_ids(csv_path: str, index_map: Dict[str, Tuple[str, str, int]], report: GroupReport) -> MappingSource:
    """
    读取 CSV, 按 index_map 将 today 的 id → job_type。
    找不到映射的, 记为 unmapped_ids; 空值记 invalid_ids; 映射不到 job_type 的给空字符串。
    """
    df = pd.read_csv(csv_path, header=None, names=["line"], delimiter=r'\n', engine="python")
    raw_ids = np.squeeze(df.values).tolist()
    today_ids = [clean_id(x) for x in raw_ids]

    pattern = re.compile(
        r'(?<!\w)(' + '|'.join(re.escape(k) for k in index_map.keys()) + r')(?!\w)'
    )

    id_list = []
    job_list = []
    job_type_list = []
    order_weights = []
    for pid in today_ids:
        # if not pid:
        #     report.invalid_ids.append(pid)
        #     continue
        # if pid in seen:
        #     report.invalid_ids.append(pid)
        #     report.add_warning(f"duplicated id in CSV: {pid}")
        #     continue

        # seen.add(pid)
        
        if "单挂" in pid:
            id_list.append(pid)
            job_list.append("单挂")
            job_type_list.append("单挂")
            order_weights.append(10000) # append a big value
            continue

        if "护肩经验" in pid:
            id_list.append(pid)
            job_list.append("单挂")
            job_type_list.append("单挂")
            order_weights.append(10000) # append a big value
            continue

        if "混次数" in pid:
            id_list.append(pid)
            job_list.append("单挂")
            job_type_list.append("单挂")
            order_weights.append(10000) # append a big value
            continue

        repo_id = list(index_map.keys())
        if pattern.findall(pid): # 精准匹配
            id_match = pattern.findall(pid)
            
            if len(id_match) == 1:
                id_name = id_match[0]
            else:
                report.unmapped_ids.append(pid)

                id_list.append(pid)
                job_list.append("未知")
                job_type_list.append("未知")
                order_weights.append(10000) # append a big value
                continue
        else:
            index_list = find_contain_index(repo_id, pid) # 模糊匹配
            if index_list and len(index_list) == 1:
                id_name = repo_id[index_list[0]]
            else:
                report.unmapped_ids.append(pid)

                id_list.append(pid)
                job_list.append("未知")
                job_type_list.append("未知")
                order_weights.append(10000) # append a big value
                continue

        if id_name in id_list:
            bug_index = id_list.index(id_name)
            id_list[bug_index] = today_ids[bug_index]
            job_list[bug_index] = "未知"
            job_type_list[bug_index] = "未知"
            order_weights[bug_index] = 10000

            report.unmapped_ids.append(pid)
            report.add_warning(f"duplicated id in CSV: [{pid}, {today_ids[bug_index]}]")
            
            id_list.append(pid)
            job_list.append("未知")
            job_type_list.append("未知")
            order_weights.append(10000) # append a big value
            continue
        
        job, job_type, weights = index_map[id_name]
        id_list.append(id_name)
        job_list.append(job)
        order_weights.append(weights)
        job_type_list.append(job_type)

    # 人数校验（可选）
    if len(today_ids) != df.shape[0]:
        report.add_warning("CSV squeeze produced different length; check data shape.")
    
    today_ms = MappingSource(order_weights, id_list, job_list, job_type_list)
    return today_ms

def main(csv_path: str = "temp.csv", repo_path: str="repo.csv", sort_by_repo_weight: bool=False):
    index_map, ms_data = build_index(repo_path)

    report = GroupReport()
    today_ms = map_today_ids(csv_path, index_map, report)
    if sort_by_repo_weight:
        today_ms.sort_by_weight()

    # today_maps: Dict[str, str] = {}
    # for i, id_i in enumerate(today_ms.ids):
    #     today_maps[id_i] = today_ms.job_types[i]

    # 问题汇总
    print("Errors:", report.errors)
    print("Unmapped:", report.unmapped_ids)
    print("Invalid:", report.invalid_ids)
    print("Leftover:", report.leftover_buckets)

    today_members: Dict[str, List[str]] = {}
    for job_i in np.unique(list(today_ms.jobs)):
        key = str(job_i)
        today_members[key] = get_key_by_value(today_ms.ids, today_ms.jobs, job_i)

    return today_members, today_ms, report

from openpyxl import load_workbook
from datetime import datetime
import string
from tabulate import tabulate

import argparse
parser = argparse.ArgumentParser(description="读取数据文件（默认 temp.csv）")
parser.add_argument(
    "file",
    nargs="?",              # 表示这个参数可有可无
    default="temp.csv",     # 如果没提供，就用默认值
    help="要读取的文件名（默认：temp.csv）"
)

if __name__ == "__main__":
    args = parser.parse_args()
    today_members, today_ms, report = main(args.file)
    for warning in report.warnings:
        print("Warnings:", warning)
    
    letters = string.ascii_uppercase
    job_order = ["奶", "火", '圣骑', '拳', '弩', '船', '饺子', '刀']

    # 1) 载入模板
    wb = load_workbook("一条分组.xlsx")
    ws = wb["Sheet1"]  # 替换为你的工作表名

    # 2) 写入文件
    now = datetime.now()

    last_index = len(job_order)-1
    for job, id_list in today_members.items():
        if job == "单挂":
            continue

        if job == "未知":
            continue
        
        if job in job_order:
            i = job_order.index(job)
        else:
            last_index += 1
            i = last_index

        ws[f'{letters[i]}1'].value = job

        for id_i, id in enumerate(id_list):
            ws[f'{letters[i]}{id_i+2}'].value = id

    job = "单挂"
    ws[f'{letters[last_index+1]}1'].value = job
    id_list = today_members[job]
    for id_i, id in enumerate(id_list):
        ws[f'{letters[last_index+1]}{id_i+2}'].value = id
    last_index += 1

    if "未知" in today_members:
        job = "未知"
        ws[f'{letters[last_index+1]}1'].value = job
        id_list = today_members[job]
        for id_i, id in enumerate(id_list):
            ws[f'{letters[last_index+1]}{id_i+2}'].value = id

    wb.save(f"{now.strftime('%Y%m%d')}一条.xlsx")
    
    # print
    df = pd.read_excel(f"{now.strftime('%Y%m%d')}一条.xlsx", sheet_name="Sheet1", header=None)
    df = df.fillna("")
    num_members = (df != "").sum().sum()-df.shape[1]
    
    id_weight = today_ms.get_weight_by_id(id)
    
    # check number of members
    df_raw = pd.read_csv(args.file, header=None, names=["line"], delimiter=r'\n', engine="python")
    if num_members != df_raw.shape[0]:
        Warning("Final excel sheet produced different length; check data shape.")

    # # print id and weight
    # for index, row in df[1:].iterrows():
    #     for col in df.columns:
    #         id_i = df.at[index, col]
    #         if id_i:
    #             id_weight = today_ms.get_weight_by_id(id_i)
    #             df.at[index, col] = f"{id_i}({id_weight})"
    #         else:
    #             continue

    print("====================")
    print(f"Total Number of Members: {num_members}")
    print(tabulate(df[1:], headers=df.iloc[0,:].values.astype('str'), tablefmt="github"))
