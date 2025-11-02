# %%
import random
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Iterable, Optional
import copy
from run import main
import pandas as pd
import numpy as np
from utils import flatten_list

import re
from typing import List

def sort_warnings(warnings: List[str]) -> List[str]:
    """
    按照 Group编号 和 Team编号 从小到大排序警告列表。

    参数:
        warnings: 包含字符串的列表，例如 ["Group2Team1 缺 奶", "Group1Team3 人数不足"]

    返回:
        排好序的警告字符串列表。
    """
    def sort_key(msg: str):
        # 从字符串中提取 Group 和 Team 的数字
        m = re.search(r'Group(\d+)Team(\d+)', msg)
        if m:
            g, t = map(int, m.groups())
            return (g, t)
        else:
            # 如果找不到，就放在最后
            return (9999, 9999)
    
    return sorted(warnings, key=sort_key)

# %%
@dataclass
class TeamSpec:
    # 用声明式“配方”定义每队所需角色，以及如何兜底
    # items 是按顺序的需求：("奶", 1) 表示抽取1人奶；("近战", 3) 表示抽3个近战
    main_character: str
    fallback: Optional[List[str]] = None  # 可选：当主要类型空时，用次选类型补齐

def pop_random(lst: List, rng: random.Random) -> Optional[str]:
    if not lst:
        return None
    i = rng.randrange(len(lst))
    lst[i], lst[-1] = lst[-1], lst[i]
    return lst.pop()

def pop_order(lst: List) -> Optional[str]:
    if not lst:
        return None
    return lst.pop(0)

def form_team(spec: TeamSpec, buckets: Dict[str, List[str]], rng: random.Random, random: bool=False) -> Tuple[List[str], List[str], bool]:
    team_ids: List[str] = []
    team_jobs_type = []

    def take(role: str, count: int) -> int:
        taken = 0
        for _ in range(count):
            if random:
                pid = pop_random(buckets.get(role, []), rng)
            else:
                pid = pop_order(buckets.get(role, []))

            if pid is None:
                break
            team_ids.append(pid)
            taken += 1
        return taken

    # 先按主需求取
    role = spec.main_character[0]
    n = spec.main_character[1]
    got = take(role, n)
    team_jobs_type.append([role]*got)

    # 不够的从备选职业取
    if spec.fallback:
        for fallback_i in spec.fallback:
            if got < n:
                fb_role = fallback_i
                fb_got = take(fb_role, n-got)
                got += fb_got
                team_jobs_type.append([fb_role]*fb_got)

    # check number of member
    if got != n:
        full = False
    else:
        full = True

    return team_ids, team_jobs_type, full

def build_teams(roles_all, numbers_all, team_number, today_map, report, rng, num_member):
    """
    根据角色分配信息和权重表 today_map 组队。

    参数:
        roles_all: List[List[角色或角色列表]]   # 每个团队的角色信息
        numbers_all: List[List[int]]            # 每个团队对应的编号信息
        today_map: dict                         # 权重或匹配配置
        rng: Random-like 对象，用于 form_team()
        report: 报告对象，需有 add_warning() 方法
        num_member: int                         # 每队成员数量上限

    返回:
        (team_flatten, job_flatten)
        - team_flatten: 每个队伍合并后的队员ID列表
        - job_flatten: 每个队伍合并后的职业类型列表
    """
    team_flatten = []
    job_flatten = []

    # 深拷贝保证 today_map 不被修改
    buckets = copy.deepcopy(today_map)

    # 遍历每支队伍
    for i, (roles, numbers) in enumerate(zip(roles_all, numbers_all)):
        team_i, job_i = [], []

        # 遍历当前队伍所有角色
        for role, number in zip(roles, numbers):
            # 角色可能是 ["主职业","副职业"] 这种结构
            if isinstance(role, list):
                spec = TeamSpec(main_character=(role[0], number),
                                fallback=role[1:])
            else:
                spec = TeamSpec(main_character=(role, number))

            # 生成队伍分配结果
            team_ids, team_jobs_type, full = form_team(spec, buckets, rng)

            # 如果成功组出角色，记录下来
            if team_ids:
                team_i.append(team_ids)
                job_i.append(team_jobs_type)

            # 如果资源（职业槽）没有填满
            if not full:
                which_team = team_number[i]
                report.add_warning(f"Group{which_team[0]}Team{which_team[1]} 缺 {role}")

        # 检查是否达标人数
        if (len(flatten_list(team_i)) < num_member) & (len(flatten_list(team_i)) > 0):
            which_team = team_number[i]
            report.add_warning(f"Group{which_team[0]}Team{which_team[1]} 人数不足")

        # 扁平化保存结果
        team_flatten.append(flatten_list(team_i))
        job_flatten.append(flatten_list(job_i))

    return team_flatten, job_flatten, buckets

import string
from datetime import datetime
from openpyxl import load_workbook
from tabulate import tabulate
if __name__ == "__main__":
    csv_name = "temp.csv"
    today_map, today_ms, report = main(csv_name, sort_by_repo_weight=True)
    
    ## form teams
    random_seed = 2025
    rng = random.Random(random_seed)
    num_member = 6

    # 1) 第一队:      奶1 + 火1 + 拳1 + 圣骑1 + 饺子1 + 需要拳的职业(弩，船)
    # 2) 第二队(远程): 奶1 + 眼1 + (优先远程4, 不够用眼补齐至4)
    # 3) 第三队(洗澡): 奶1 + 火1 + 刀
    team_flatten = []
    job_flatten = []
    
    roles_all = [["奶", "火", "拳", "圣骑", "饺子", ["火毒", "饺子", "圣骑", "冰雷", "拳"]],
                ["奶", "火", "拳", "圣骑", "饺子", ["冰雷", "饺子", "圣骑", "火毒", "拳"]],
                ["奶", "火", ["刀", "饺子"]],
                ["奶", "火", ["刀", "饺子"]],
                ["奶", "火", "弓", ["标", "弓", "船", "弩"]],
                ["奶", "火", "弓", ["标", "弓", "弩", "船"]],]       
    numbers_all = [[1, 1, 1, 1, 1, 1],
                [1, 1, 1, 1, 1, 1],
                [1, 1, 4],
                [1, 1, 4],
                [1, 1, 1, 3],
                [1, 1, 1, 3],]
    team_number = [[1,1],[2,1],[1,3],[2,3],[1,2],[2,2]]

    team_flatten, job_flatten, buckets = build_teams(roles_all, numbers_all, team_number, today_map, report, rng, num_member)

    # 3) 第四队(剩余人数按远程近战分组)
    id_remain = []
    job_remain = []
    job_type_remain = []
    for k, v in buckets.items(): 
        if k != '单挂' and v:
            id_remain.extend(v)
            job_remain.extend([k]*len(v))
            job_type_remain.extend([today_ms.get_jtype_by_id(v_i) for v_i in v])

    remain_map: Dict[str, str] = {}
    for i,id_i in enumerate(id_remain):
        remain_map[id_i] = job_remain[i]

    buckets_remain: Dict[str, List[str]] = {}
    for job_i in set(job_type_remain):
        for i,v in enumerate(job_type_remain):
            if v==job_i:
                if job_type_remain[i] not in buckets_remain:
                    buckets_remain[job_type_remain[i]] = []
                buckets_remain[job_type_remain[i]].append(id_remain[i])

    roles_all = [['奶', ['近战','远程']],['奶', ['远程','近战']]]
    numbers_all = [[1,5], [1,5]]
    team_number_2 = [[1,4],[2,4]]

    team_remain, _, buckets_remain_remain = build_teams(roles_all, numbers_all, team_number_2, buckets_remain, report, rng, num_member)
    team_flatten.extend(team_remain)
    for id_team in team_remain:
        job_flatten.append([remain_map[id_i] for id_i in id_team])
    team_number.extend(team_number_2)
    
    # print warning
    for warning in sort_warnings(report.warnings):
        print("Warnings:", warning)

    print(buckets_remain_remain)

    ## 
    letters = string.ascii_uppercase

    # 1) 载入模板
    wb = load_workbook("一条排班.xlsx")
    ws = wb["Sheet1"]  # 替换为你的工作表名

    # 2) 写入文件
    now = datetime.now()

    for i, (job, id_list) in enumerate(zip(job_flatten, team_flatten)):
        for j, job_i in enumerate(job):
            
            team_number_i = team_number[i]
            if team_number_i[0] == 1:
                i_mod = team_number_i[1]-1
                ws[f'{letters[i_mod*2]}{3+j}'].value = id_list[j]
                ws[f'{letters[i_mod*2+1]}{3+j}'].value = job_i
            else:
                i_mod = team_number_i[1]-1
                ws[f'{letters[i_mod*2]}{12+j}'].value = id_list[j]
                ws[f'{letters[i_mod*2+1]}{12+j}'].value = job_i

    ws['k2'] = '单挂'
    for i, id_i in enumerate(buckets['单挂']):
        ws[f'k{2+i+1}'] = id_i
    
    i += 2
    for key, value in buckets_remain_remain.items():
        for v_i in value:
            id_i = f"{key}: {v_i}"
            ws[f'k{2+i+1}'] = id_i
            i += 1

    wb.save(f"{now.strftime('%Y%m%d')}一条排班.xlsx")


    # print
    df = pd.read_excel(f"{now.strftime('%Y%m%d')}一条排班.xlsx", sheet_name="Sheet1", header=None)
    df = df.fillna("")

    print("====================")
    print(tabulate(df[1:], headers=df.iloc[0,:].values.astype('str'), tablefmt="github"))
